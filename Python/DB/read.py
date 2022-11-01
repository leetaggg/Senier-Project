from openpyxl import load_workbook, Workbook
import pandas as pd
import mysql.connector
from mysql.connector import Error

def user_down(path):
    try:
        connection = mysql.connector.connect(host='39.124.26.132',
                                                database='student',
                                                user='root',
                                                password='123456')

        cursor = connection.cursor()
        sql = "SELECT * FROM attend;"

        cursor.execute(sql, )
        record = cursor.fetchall()
        for i in range(0, len(record)):
            record[i] = list(record[i])
            for e in range(1, len(record[i])):
                if record[i][e] == None:
                    record[i][e] = 'None'
                    
    except mysql.connector.Error as error:
        print("연결 실패 {}".format(error))
        return False

    finally:
        if (connection.is_connected()):
            cursor.close()
            connection.close()
            print("출석 명단 저장 완료")
            
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws.append(['학번', '1주차', '2주차', '3주차', '4주차', '5주차'])
    for row in record:
        row = list(row)
        write_ws.append(row)
    write_wb.save(path)
    return True

def user_upload(path):
    week_nbls = {0:"zero", 1:"one", 2:"two", 3:"three", 4:"four", 5:"five", 6:"six", 7:"seven", 8:"eight", 9:"nine", 10:"ten"}
    try:
        connection = mysql.connector.connect(host='39.124.26.132',
                                                database='student',
                                                user='root',
                                                password='123456')

        cursor = connection.cursor(prepared = True)
        
        load_wb = load_workbook(path, data_only=True)
        load_ws = load_wb['Sheet']
        all_values = []
        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
            
        del all_values[0]

        for i in range(0, len(all_values)):
            all_values[i] = list(all_values[i])
            for e in range(1, len(all_values[i])):
                if all_values[i][e] == 'None':
                    all_values[i][e] = None

        for i in range(0, len(all_values)):
            for e in range(0, len(all_values[i])-1):
                week_number = week_nbls[e+1]
                week_number = "week"+week_number
                atd = all_values[i][e+1]
                std_id = all_values[i][0]
                sql = 'UPDATE attend SET {0} = %s WHERE userID = %s;'.format(week_number)
                data = (atd, std_id)
                cursor.execute(sql, data)
                connection.commit()

    except mysql.connector.Error as error:
        print("업로드 실패 {}".format(error))
        return False

    finally:
        if (connection.is_connected()):
            cursor.close()
            connection.close()
            print("출석 명단 업로드 성공")
            return True