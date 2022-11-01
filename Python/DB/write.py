from openpyxl import Workbook
import mysql.connector
from mysql.connector import Error

'''write_wb = Workbook()

write_ws = write_wb.create_sheet('생성시트')

write_ws = write_wb.active
write_ws['A1'] = '숫자'

write_ws.append([1, 2, 3])

write_ws.cell(5, 5, '5행5열')
write_wb.save("D:/VS Code/mysql/helloworld.xlsx")'''

def user_down():
    try:
        connection = mysql.connector.connect(host='39.124.26.132',
                                                database='student',
                                                user='root',
                                                password='123456')

        cursor = connection.cursor()
        sql = "SELECT * FROM attend;"

        cursor.execute(sql, )
        record = cursor.fetchall()
        for i in range(0, len(record)):
            record[i] = list(record[i])
            for e in range(1, len(record[i])):
                if record[i][e] == None:
                    record[i][e] = 'None'
                    
    except mysql.connector.Error as error:
        print("연결 실패 {}".format(error))

    finally:
        if (connection.is_connected()):
            cursor.close()
            connection.close()
            print("접속 종료")
            
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws.append(['학번', '1주차', '2주차', '3주차', '4주차', '5주차'])
    for row in record:
        row = list(row)
        write_ws.append(row)
    write_wb.save("C:/AHard/Project/DB/attendance.xlsx")

